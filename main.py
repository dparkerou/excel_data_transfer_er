#!/usr/bin/env python3
import openpyxl
from openpyxl.utils import get_column_letter


def main():
    source_worksheet_name = ""
    destination_worksheet_name = ""
    
    #import the source csv
    path_to_source_spreadsheet = input("Drag and drop the source file here: ").rstrip()
    source_spreadsheet_object = openpyxl.load_workbook(filename = path_to_source_spreadsheet, data_only = True)
    source_active_sheet_object = source_spreadsheet_object[source_worksheet_name] #change this to correct sheet
    
    #import the dest csv
    path_to_dest_spreadsheet = input("Please drag and drop the destination file here: ").rstrip()
    destination_spreadsheet_object = openpyxl.load_workbook(filename = path_to_dest_spreadsheet, data_only = True)
    destination_active_sheet_object = destination_spreadsheet_object[destination_worksheet_name]
    
    #define the parameters of the data
    start = input("Which cell do you want to start with? ").rstrip().upper()
    finish = input("Which cell do you want to end with? ").rstrip().upper()
    source_data_we_want = source_active_sheet_object[start : finish]

    #get the list of names from the destination spreadsheet
    dest_names = []
    row = destination_active_sheet_object.max_row
    for i in range(1, row + 1):
        cell_obj = destination_active_sheet_object.cell(row = i, column = 1)
        dest_names.append(cell_obj.value)

    #get employee data from the source
    for emp_name, program, title in source_data_we_want:
        if emp_name.value not in dest_names:
            employee_to_add = [emp_name.value, program.value, title.value]
            destination_active_sheet_object.append(employee_to_add)
            print("Records added to document: ")
            print(emp_name.value, program.value, title.value)
    destination_spreadsheet_object.save(filename=path_to_dest_spreadsheet)

if __name__ == "__main__":
    main()
